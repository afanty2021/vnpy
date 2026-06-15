"""
港股通股票名单爬虫

从深圳证券交易所和上海证券交易所下载港股通标的名单Excel文件。
深交所和上交所的港股通名单是相同的，优先使用深交所API。
"""

import logging
import io
from typing import List, Optional
from datetime import date

try:
    import requests
    import openpyxl
    HAS_DEPS = True
except ImportError:
    HAS_DEPS = False

from ..models.hk_connect import HkConnectStock


logger = logging.getLogger("vnpy_china_data")


class HkConnectCrawler:
    """港股通股票名单爬虫

    从深圳证券交易所和上海证券交易所下载港股通标的名单Excel文件。
    深交所和上交所的港股通名单是相同的，优先使用深交所API，
    上交所API作为备用。
    """

    # 港股通股票名单数据源
    # 深交所和上交所的港股通名单是相同的，使用深交所API为主

    # 深交所深港通Excel API（主要数据源）
    # 格式：证券代码、中文简称、英文简称
    SZSE_HK_CONNECT_XLSX_URL = "https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=SGT_GGTBDQD&TABKEY=tab1"

    # 上交所沪港通Excel API（备用数据源）
    SSE_HK_CONNECT_XLSX_URL = "https://query.sse.com.cn/commonExcelDd.do?sqlId=COMMON_SSE_JYFW_HGT_XXPL_BDZQQD_L&keyword="

    def __init__(self, timeout: int = 30):
        """初始化爬虫

        Args:
            timeout: 请求超时时间（秒）
        """
        self.timeout = timeout
        self.session = None

        if not HAS_DEPS:
            logger.warning(
                "港股通爬虫需要安装 requests 和 openpyxl："
                "pip install requests openpyxl"
            )

    def crawl_all(self) -> List[HkConnectStock]:
        """爬取所有港股通股票名单

        深交所和上交所的港股通名单是相同的，优先使用深交所API，
        上交所API作为备用。

        Returns:
            港股通股票列表
        """
        if not HAS_DEPS:
            logger.error("缺少依赖库，无法爬取港股通名单")
            return []

        # 优先使用深交所API
        try:
            stocks = self.crawl_szse_hk_connect()
            if stocks:
                logger.info(f"从深交所爬取港股通股票 {len(stocks)} 只")
                return stocks
        except Exception as e:
            logger.warning(f"从深交所爬取失败，尝试上交所: {e}")

        # 备用：使用上交所API
        try:
            stocks = self.crawl_sse_hk_connect()
            if stocks:
                logger.info(f"从上交所爬取港股通股票 {len(stocks)} 只")
                return stocks
        except Exception as e:
            logger.error(f"从上交所爬取失败: {e}")

        logger.warning("无法从任何数据源爬取港股通名单")
        return []

    def crawl_sse_hk_connect(self) -> List[HkConnectStock]:
        """爬取上交所沪港通股票名单

        从上交所Excel文件中获取沪港通港股名单。

        Returns:
            沪港通股票列表
        """
        if not HAS_DEPS:
            return []

        stocks = []

        try:
            # 下载Excel文件
            response = requests.get(
                self.SSE_HK_CONNECT_XLSX_URL,
                timeout=self.timeout,
                headers={"User-Agent": "Mozilla/5.0"}
            )
            response.raise_for_status()

            # 解析Excel
            wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            ws = wb.active

            # 跳过表头（第一行）
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 格式：证券代码、中文简称、英文简称
                if len(row) >= 2 and row[0] and row[1]:
                    stock_code = str(row[0]).strip()
                    stock_name = str(row[1]).strip()

                    # 港股代码通常是4-5位数字
                    if stock_code.isdigit() and 1 <= len(stock_code) <= 5:
                        stock = HkConnectStock(
                            symbol=stock_code.zfill(5),
                            name=stock_name,
                            channel="SHHK",
                            channel_type="SH",
                            category="沪港通",
                            status="active",
                            source="sse",
                        )
                        stocks.append(stock)

        except Exception as e:
            logger.error(f"爬取上交所沪港通名单失败: {e}")

        return stocks

    def crawl_szse_hk_connect(self) -> List[HkConnectStock]:
        """爬取深交所深港通股票名单

        从深交所Excel文件中获取深港通港股名单。

        Returns:
            深港通股票列表
        """
        if not HAS_DEPS:
            return []

        stocks = []

        try:
            # 下载Excel文件
            response = requests.get(
                self.SZSE_HK_CONNECT_XLSX_URL,
                timeout=self.timeout,
                headers={"User-Agent": "Mozilla/5.0"}
            )
            response.raise_for_status()

            # 解析Excel
            wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            ws = wb.active

            # 跳过表头（第一行）
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 格式：证券代码、中文简称、英文简称
                if len(row) >= 2 and row[0] and row[1]:
                    stock_code = str(row[0]).strip()
                    stock_name = str(row[1]).strip()

                    # 港股代码通常是4-5位数字
                    if stock_code.isdigit() and 1 <= len(stock_code) <= 5:
                        stock = HkConnectStock(
                            symbol=stock_code.zfill(5),
                            name=stock_name,
                            channel="SZHK",
                            channel_type="SZ",
                            category="深港通",
                            status="active",
                            source="szse",
                        )
                        stocks.append(stock)

        except Exception as e:
            logger.error(f"爬取深交所深港通名单失败: {e}")

        return stocks


# 便捷函数
def crawl_hk_connect_stocks() -> List[HkConnectStock]:
    """爬取港股通股票名单的便捷函数

    Returns:
        港股通股票列表
    """
    crawler = HkConnectCrawler()
    return crawler.crawl_all()
