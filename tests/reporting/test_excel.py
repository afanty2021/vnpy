"""
Excel导出器单元测试
"""

import pytest
import os
import tempfile
from datetime import date, datetime
from pathlib import Path

from vnpy_china_reporting.core.models import (
    ReportData, PositionAnalysis, PositionRecord, AccountData
)
from vnpy_china_reporting.core.enums import ReportType, PositionSide
from vnpy_china_reporting.export import ExcelExporter


class TestExcelExporter:
    """Excel导出器测试类"""

    @pytest.fixture
    def exporter(self):
        """创建导出器实例"""
        return ExcelExporter()

    @pytest.fixture
    def sample_report(self):
        """创建示例报表数据"""
        # 创建账户数据
        account = AccountData(
            total_equity=1000000.0,
            available_cash=500000.0,
            market_value=500000.0,
            total_pnl=50000.0,
            total_pnl_ratio=0.05,
            commission=1000.0,
            timestamp=datetime.now()
        )

        # 创建持仓数据
        positions = [
            PositionRecord(
                symbol="000001",
                name="平安银行",
                side=PositionSide.LONG,
                volume=10000,
                avg_cost=10.5,
                current_price=11.0,
                market_value=110000.0,
                unrealized_pnl=5000.0,
                unrealized_pnl_ratio=0.0476
            ),
            PositionRecord(
                symbol="600000",
                name="浦发银行",
                side=PositionSide.LONG,
                volume=20000,
                avg_cost=8.0,
                current_price=7.5,
                market_value=150000.0,
                unrealized_pnl=-10000.0,
                unrealized_pnl_ratio=-0.0625
            ),
        ]

        # 创建交易数据
        from vnpy_china_reporting.core.models import TradeRecord
        trades = [
            TradeRecord(
                trade_id="T001",
                symbol="000001",
                direction="buy",
                volume=10000,
                price=10.5,
                amount=105000.0,
                commission=50.0,
                timestamp=datetime.now()
            ),
            TradeRecord(
                trade_id="T002",
                symbol="600000",
                direction="sell",
                volume=20000,
                price=7.5,
                amount=150000.0,
                commission=75.0,
                timestamp=datetime.now()
            ),
        ]

        # 创建报表数据
        report = ReportData(
            report_type=ReportType.DAILY,
            start_date=date(2026, 2, 25),
            end_date=date(2026, 2, 25),
            account=account,
            positions=positions,
            trades=trades,
            daily_pnl=50000.0,
            daily_pnl_ratio=0.05
        )

        return report

    @pytest.fixture
    def sample_analysis(self):
        """创建示例持仓分析数据"""
        analysis = PositionAnalysis(
            total_positions=5,
            total_market_value=1000000.0,
            top_holdings=[
                {
                    "symbol": "000001",
                    "name": "平安银行",
                    "market_value": 300000.0,
                    "weight": 0.3,
                    "pnl": 10000.0,
                    "pnl_ratio": 0.05
                },
                {
                    "symbol": "600000",
                    "name": "浦发银行",
                    "market_value": 200000.0,
                    "weight": 0.2,
                    "pnl": -5000.0,
                    "pnl_ratio": -0.025
                },
            ],
            concentration=0.5,
            industry_distribution={
                "银行": {
                    "value": 500000.0,
                    "ratio": 0.5,
                    "count": 2,
                    "avg_pnl": 2500.0
                },
                "地产": {
                    "value": 500000.0,
                    "ratio": 0.5,
                    "count": 3,
                    "avg_pnl": 1000.0
                }
            }
        )
        return analysis

    def test_exporter_init(self, exporter):
        """测试导出器初始化"""
        assert exporter is not None
        assert exporter.title_font is not None
        assert exporter.header_font is not None
        assert exporter.normal_font is not None
        assert exporter.title_fill is not None
        assert exporter.header_fill is not None

    def test_export_daily_report(self, exporter, sample_report):
        """测试导出日报"""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_daily.xlsx")

            result = exporter.export_daily_report(sample_report, filepath)

            assert result is True
            assert os.path.exists(filepath)
            assert os.path.getsize(filepath) > 0

    def test_export_monthly_report(self, exporter, sample_report):
        """测试导出月报"""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_monthly.xlsx")

            result = exporter.export_monthly_report(sample_report, filepath)

            assert result is True
            assert os.path.exists(filepath)

    def test_export_position_analysis(self, exporter, sample_analysis):
        """测试导出持仓分析"""
        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_analysis.xlsx")

            result = exporter.export_position_analysis(sample_analysis, filepath)

            assert result is True
            assert os.path.exists(filepath)

    def test_export_with_empty_positions(self, exporter, sample_report):
        """测试空持仓情况"""
        sample_report.positions = []

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_empty.xlsx")

            result = exporter.export_daily_report(sample_report, filepath)

            assert result is True
            assert os.path.exists(filepath)

    def test_export_with_empty_trades(self, exporter, sample_report):
        """测试无交易情况"""
        sample_report.trades = []

        with tempfile.TemporaryDirectory() as tmpdir:
            filepath = os.path.join(tmpdir, "test_no_trades.xlsx")

            result = exporter.export_daily_report(sample_report, filepath)

            assert result is True
            assert os.path.exists(filepath)

    def test_write_title(self, exporter):
        """测试标题写入"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        row = exporter._write_title(ws, 1, "测试标题")

        assert ws.cell(row=1, column=1).value == "测试标题"

    def test_write_section(self, exporter):
        """测试区块写入"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        data = [
            ["字段1", "值1"],
            ["字段2", "值2"]
        ]

        row = exporter._write_section(ws, 1, "测试区块", data)

        assert ws.cell(row=1, column=1).value == "测试区块"
        assert ws.cell(row=2, column=1).value == "字段1"
        assert ws.cell(row=2, column=2).value == "值1"

    def test_apply_styles(self, exporter):
        """测试样式应用"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "测试"

        exporter.apply_styles(ws)

        assert ws['A1'].font is not None
