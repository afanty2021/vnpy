"""
Excel导出器

将报表数据导出为Excel格式。
"""

from __future__ import annotations

from typing import Optional, List, Any
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, datetime

from ..core.models import ReportData, PositionAnalysis, PositionRecord


class ExcelExporter:
    """
    Excel导出器

    将报表数据导出为Excel文件，支持日报、月报、持仓分析等格式。
    """

    def __init__(self) -> None:
        """初始化导出器并设置默认样式"""
        # 字体设置
        self.title_font = Font(
            size=14, bold=True, color="FFFFFF", name="微软雅黑"
        )
        self.header_font = Font(
            size=11, bold=True, color="333333", name="微软雅黑"
        )
        self.normal_font = Font(size=10, name="微软雅黑")

        # 填充颜色
        self.title_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self.header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        # 对齐方式
        self.center_align = Alignment(
            horizontal="center", vertical="center"
        )
        self.left_align = Alignment(
            horizontal="left", vertical="center"
        )
        self.right_align = Alignment(
            horizontal="right", vertical="center"
        )

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.border = thin_border

    def export_daily_report(
        self,
        report: ReportData,
        filepath: str
    ) -> bool:
        """
        导出日报为Excel文件

        Args:
            report: 报表数据对象
            filepath: 导出文件路径

        Returns:
            是否导出成功
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "日报"

            row = 1

            # 写入标题
            title_text = f"每日交易报告 - {report.start_date}"
            row = self._write_title(ws, row, title_text)
            row += 1

            # 资金情况
            if report.account:
                row = self._write_section(
                    ws, row, "资金情况", [
                        ["总权益", f"{report.account.total_equity:.2f}"],
                        ["可用资金", f"{report.account.available_cash:.2f}"],
                        ["持仓市值", f"{report.account.market_value:.2f}"],
                        ["总盈亏", f"{report.account.total_pnl:.2f}"],
                        ["盈亏比例", f"{report.account.total_pnl_ratio:.2%}"]
                    ]
                )

            # 当期盈亏
            row = self._write_section(
                ws, row, "当日盈亏", [
                    ["当日盈亏", f"{report.daily_pnl:.2f}" if report.daily_pnl is not None else "N/A"],
                    ["盈亏比例", f"{report.daily_pnl_ratio:.2%}" if report.daily_pnl_ratio is not None else "N/A"]
                ]
            )

            # 交易统计
            if report.trades:
                buy_count = sum(1 for t in report.trades if t.direction == "buy")
                sell_count = sum(1 for t in report.trades if t.direction == "sell")
                total_amount = sum(t.amount for t in report.trades)
                total_commission = sum(t.commission for t in report.trades)

                row = self._write_section(
                    ws, row, "交易统计", [
                        ["总成交笔数", len(report.trades)],
                        ["买入笔数", buy_count],
                        ["卖出笔数", sell_count],
                        ["总成交金额", f"{total_amount:.2f}"],
                        ["总手续费", f"{total_commission:.2f}"]
                    ]
                )

            # 持仓明细
            if report.positions:
                row = self._write_position_table(ws, row, report.positions)

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"导出日报失败: {e}")
            return False

    def export_monthly_report(
        self,
        report: ReportData,
        filepath: str
    ) -> bool:
        """
        导出月报为Excel文件

        Args:
            report: 报表数据对象
            filepath: 导出文件路径

        Returns:
            是否导出成功
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "月报"

            row = 1

            # 标题
            month_text = f"{report.start_date.year}年{report.start_date.month:02d}月交易报告"
            row = self._write_title(ws, row, month_text)
            row += 1

            # 资金情况
            if report.account:
                row = self._write_section(
                    ws, row, "账户概况", [
                        ["期初资金", f"{report.account.available_cash + report.account.market_value:.2f}"],
                        ["期末资金", f"{report.account.total_equity:.2f}"],
                        ["可用资金", f"{report.account.available_cash:.2f}"],
                        ["持仓市值", f"{report.account.market_value:.2f}"]
                    ]
                )

            # 月度盈亏
            acc = report.account
            row = self._write_section(
                ws, row, "月度盈亏", [
                    ["月度盈亏", f"{report.daily_pnl:.2f}" if report.daily_pnl is not None else "N/A"],
                    ["盈亏比例", f"{report.daily_pnl_ratio:.2%}" if report.daily_pnl_ratio is not None else "N/A"],
                    ["手续费", f"{acc.commission:.2f}" if acc else "N/A"]
                ]
            )

            # 交易统计
            if report.trades:
                row = self._write_section(
                    ws, row, "交易统计", [
                        ["总成交笔数", len(report.trades)],
                        ["成交股票数", len(set(t.symbol for t in report.trades))]
                    ]
                )

            # 持仓明细
            if report.positions:
                row = self._write_position_table(ws, row, report.positions)

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"导出月报失败: {e}")
            return False

    def export_position_analysis(
        self,
        analysis: PositionAnalysis,
        filepath: str
    ) -> bool:
        """
        导出持仓分析为Excel文件

        Args:
            analysis: 持仓分析数据对象
            filepath: 导出文件路径

        Returns:
            是否导出成功
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "持仓分析"

            row = 1

            # 标题
            row = self._write_title(ws, row, "持仓分析报告")
            row += 1

            # 总体统计
            row = self._write_section(
                ws, row, "总体统计", [
                    ["持仓数量", analysis.total_positions],
                    ["总市值", f"{analysis.total_market_value:.2f}"],
                    ["集中度", f"{analysis.concentration:.2%}"]
                ]
            )

            # 重点持仓
            if analysis.top_holdings:
                row = self._write_top_holdings_table(ws, row, analysis.top_holdings)

            # 行业分布
            if analysis.industry_distribution:
                row = self._write_industry_table(
                    ws, row, analysis.industry_distribution
                )

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"导出持仓分析失败: {e}")
            return False

    def _write_title(self, ws: Worksheet, row: int, text: str) -> int:
        """
        写入标题行

        Args:
            ws: 工作表对象
            row: 行号
            text: 标题文本

        Returns:
            下一行行号
        """
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        ws.row_dimensions[row].height = 25
        return row + 1

    def _write_section(
        self,
        ws: Worksheet,
        row: int,
        title: str,
        data: List[List[Any]]
    ) -> int:
        """
        写入一个数据区块

        Args:
            ws: 工作表对象
            row: 行号
            title: 区块标题
            data: 数据列表 [[label, value], ...]

        Returns:
            下一行行号
        """
        # 区块标题
        ws.cell(row=row, column=1, value=title).font = self.header_font
        row += 1

        # 数据行
        for label, value in data:
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_label.alignment = self.left_align

            cell_value = ws.cell(row=row, column=2, value=value)
            cell_value.alignment = self.right_align

            row += 1

        return row + 1

    def _write_position_table(
        self,
        ws: Worksheet,
        row: int,
        positions: List[PositionRecord]
    ) -> int:
        """
        写入持仓明细表

        Args:
            ws: 工作表对象
            row: 行号
            positions: 持仓列表

        Returns:
            下一行行号
        """
        # 表标题
        ws.cell(row=row, column=1, value="持仓明细").font = self.header_font
        row += 1

        # 表头
        headers = ["代码", "名称", "数量", "成本价", "现价", "市值", "盈亏", "盈亏比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        # 数据行
        for pos in positions:
            data = [
                pos.symbol,
                pos.name,
                pos.volume,
                f"{pos.avg_cost:.2f}",
                f"{pos.current_price:.2f}",
                f"{pos.market_value:.2f}",
                f"{pos.unrealized_pnl:.2f}",
                f"{pos.unrealized_pnl_ratio:.2%}"
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col <= 2:
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.right_align

                # 盈亏颜色
                if col == 7:  # 盈亏列
                    if isinstance(value, str):
                        pnl_value = float(value)
                    else:
                        pnl_value = value  # type: ignore
                    if isinstance(pnl_value, (int, float)) and pnl_value > 0:
                        cell.font = Font(color="FF0000", name="微软雅黑")  # 红色
                    elif pnl_value < 0:
                        cell.font = Font(color="00AA00", name="微软雅黑")  # 绿色

                cell.border = self.border
            row += 1

        return row + 1

    def _write_top_holdings_table(
        self,
        ws: Worksheet,
        row: int,
        top_holdings: List[dict]
    ) -> int:
        """
        写入重点持仓表

        Args:
            ws: 工作表对象
            row: 行号
            top_holdings: 重点持仓列表

        Returns:
            下一行行号
        """
        ws.cell(row=row, column=1, value="重点持仓").font = self.header_font
        row += 1

        # 表头
        headers = ["代码", "名称", "市值", "占比", "盈亏", "盈亏比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        # 数据
        for holding in top_holdings:
            data = [
                holding.get("symbol", ""),
                holding.get("name", ""),
                f"{holding.get('market_value', 0):.2f}",
                f"{holding.get('weight', 0):.2%}",
                f"{holding.get('pnl', 0):.2f}",
                f"{holding.get('pnl_ratio', 0):.2%}"
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = self.left_align if col <= 2 else self.right_align
                cell.border = self.border
            row += 1

        return row + 1

    def _write_industry_table(
        self,
        ws: Worksheet,
        row: int,
        industry_data: dict
    ) -> int:
        """
        写入行业分布表

        Args:
            ws: 工作表对象
            row: 行号
            industry_data: 行业分布数据

        Returns:
            下一行行号
        """
        ws.cell(row=row, column=1, value="行业分布").font = self.header_font
        row += 1

        # 表头
        headers = ["行业", "市值", "占比", "数量", "平均盈亏"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        # 数据
        for industry, stats in industry_data.items():
            data = [
                industry,
                f"{stats.get('value', 0):.2f}",
                f"{stats.get('ratio', 0):.2%}",
                stats.get('count', 0),
                f"{stats.get('avg_pnl', 0):.2f}"
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = self.left_align if col == 1 else self.right_align
                cell.border = self.border
            row += 1

        return row + 1

    def apply_styles(
        self,
        worksheet: Worksheet,
        styles: Optional[dict] = None
    ) -> None:
        """
        应用样式到工作表

        Args:
            worksheet: 工作表对象
            styles: 样式字典（可选）
        """
        if styles is None:
            styles = self._get_default_styles()

        for cell in worksheet['A1:Z1']:
            for c in cell:
                c.font = styles.get('header_font', self.header_font)
                c.fill = styles.get('header_fill', self.header_fill)
                c.alignment = styles.get('header_align', self.center_align)

    def _get_default_styles(self) -> dict:
        """
        获取默认样式字典

        Returns:
            样式字典
        """
        return {
            'title_font': self.title_font,
            'header_font': self.header_font,
            'normal_font': self.normal_font,
            'title_fill': self.title_fill,
            'header_fill': self.header_fill,
            'center_align': self.center_align,
            'left_align': self.left_align,
            'right_align': self.right_align,
            'border': self.border
        }
